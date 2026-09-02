import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

import regles_config

# --- RÉFÉRENCES COMMUNES ---
FP2E_REGEX = r'^[A-Z]\d{2}[A-Z]{2}\d{6}$'  # Lettre, AA, LL, 6 chiffres
FP2E_WITH_SUFFIX_REGEX = r'^[A-Z]\d{2}[A-Z]{2}\d{6}[A-Z]$'  # FP2E + 1 lettre finale
# Lettre -> diamètre(s) FP2E : valeurs par défaut (repli). La table réellement
# utilisée est lue depuis la configuration Excel (onglet Diametre_FP2E).
FP2E_DIAM_MAP = {k: list(v) for k, v in regles_config.DEFAUT_DIAMETRE_FP2E.items()}


def _diam_map_fp2e():
    """Table lettre -> [diamètres] issue de la configuration (avec repli)."""
    return regles_config.get_config().diametre_fp2e

# Repli local si la configuration Excel est indisponible (valeurs par défaut).
TYPE_COMPTEUR_VALIDES = frozenset(regles_config.DEFAUT_TYPES_COMPTEUR)


def appliquer_longueur_tete(df_with_anomalies, mode, cfg):
    """Applique les règles de longueur de tête (Mode/Marque/Type Compteur)
    définies dans la configuration Excel.

    Règle la plus spécifique (Marque + Type) prioritaire sur la règle générale
    (Marque seule). N'agit que si le numéro de tête est renseigné.
    """
    regles = cfg.regles_tete(mode)
    if not regles:
        return

    marque_norm = df_with_anomalies['Marque'].astype(str).str.upper().str.replace(' ', '', regex=False)
    type_norm = df_with_anomalies['Type Compteur'].astype(str).str.upper().str.replace(' ', '', regex=False)
    tete = df_with_anomalies['Numéro de tête'].astype(str)
    tete_renseignee = ~tete.isin(['', 'nan'])

    longueur_attendue = pd.Series([pd.NA] * len(df_with_anomalies), index=df_with_anomalies.index, dtype='object')
    message = pd.Series([''] * len(df_with_anomalies), index=df_with_anomalies.index)

    # Règles générales d'abord, spécifiques ensuite => la spécifique écrase
    for (m_norm, t_norm, longueur, msg) in regles:
        sel = tete_renseignee & (marque_norm == m_norm)
        if t_norm is not None:
            sel = sel & (type_norm == t_norm)
        longueur_attendue[sel] = longueur
        message[sel] = msg

    longueur_num = pd.to_numeric(longueur_attendue, errors='coerce')
    mauvais = longueur_num.notna() & (tete.str.len() != longueur_num)
    df_with_anomalies.loc[mauvais, 'Anomalie'] += message[mauvais] + ' / '


def colonnes_surlignage_defaut(libelle):
    """Déduit, à partir des mots-clés du libellé d'anomalie, les colonnes à
    surligner. Sert de repli lorsqu'un libellé n'est pas dans la table exacte
    (notamment pour les règles ajoutées via la configuration Excel)."""
    l = str(libelle).lower()
    cols = []
    if 'tête' in l or 'tete' in l:
        cols.append('Numéro de tête')
    if 'diamètre' in l or 'diametre' in l:
        cols.append('Diametre')
    if 'protocole' in l:
        cols.append('Protocole Radio')
    if 'marque' in l:
        cols.append('Marque')
    if 'type compteur' in l:
        cols.append('Type Compteur')
    if 'gps' in l or 'coordonnées' in l or 'coordonnees' in l:
        cols += ['Latitude', 'Longitude']
    if 'année' in l or 'annee' in l or 'millésime' in l or 'millesime' in l:
        cols.append('Année de fabrication')
    if 'compteur' in l and 'type compteur' not in l:
        cols.append('Numéro de compteur')
    return cols


# --- FONCTIONS DE VÉRIFICATION ---


def check_fp2e_details_radio(row):
    """
    Vérifie pour SAPPEL/ITRON côté radio (ou FP2E manuel toléré) :
      - cohérence année (millésime dans le n° vs 'Année de fabrication' 2 derniers digits),
      - cohérence diamètre vs lettre FP2E (4e char).
    Renvoie (liste_anomalies, dict_corrections).
    """
    anomalies, corrections = [], {}
    try:
        compteur = str(row['Numéro de compteur']).strip()
        annee_fabrication_val = str(row['Année de fabrication']).strip()
        diametre_val = row['Diametre']

        # Si le format compteur n'est pas FP2E, rien à signaler ici
        if not re.match(FP2E_REGEX, compteur):
            return [], {}

        # Millésime = positions 1-2 ; lettre diamètre = position 4
        annee_compteur = compteur[1:3]
        lettre_diam = compteur[4].upper()

        # Année : vide / non numérique / non égale au millésime => propose correction
        if (
            annee_fabrication_val == ''
            or not annee_fabrication_val.isdigit()
            or annee_compteur != annee_fabrication_val.zfill(2)
        ):
            anomalies.append("L'année de millésime n'est pas conforme")
            corrections['annee'] = annee_compteur

        # Diamètre attendu depuis la lettre
        expected_diametres = _diam_map_fp2e().get(lettre_diam, [])
        if not isinstance(expected_diametres, list):
            expected_diametres = [expected_diametres]

        if pd.isna(diametre_val) or diametre_val not in expected_diametres:
            anomalies.append("Le diamètre n'est pas conforme")
            # Correction proposée : 1er diamètre listé pour la lettre
            if expected_diametres:
                corrections['diametre'] = str(expected_diametres[0])

    except (TypeError, ValueError, IndexError):
        anomalies.append("Le numéro de compteur n'est pas conforme")

    return anomalies, corrections


def check_fp2e_details_tele(row):
    """
    Variante télée relevée : mêmes contrôles FP2E que 'radio',
    messages adaptés et mêmes propositions de corrections.
    """
    anomalies, corrections = [], {}
    try:
        compteur = str(row['Numéro de compteur']).strip()
        annee_fabrication_val = str(row['Année de fabrication']).strip()
        diametre_val = row['Diametre']

        if not re.match(FP2E_REGEX, compteur):
            anomalies.append('Format de compteur non FP2E')
            return anomalies, corrections

        annee_compteur = compteur[1:3]
        lettre_diam = compteur[4].upper()

        if (
            not annee_fabrication_val
            or not annee_fabrication_val.isdigit()
            or annee_compteur != annee_fabrication_val.zfill(2)
        ):
            anomalies.append('Année millésime non conforme FP2E')
            corrections['annee'] = annee_compteur

        expected_diametres = _diam_map_fp2e().get(lettre_diam, [])
        if not isinstance(expected_diametres, list):
            expected_diametres = [expected_diametres]

        if pd.isna(diametre_val) or diametre_val not in expected_diametres:
            anomalies.append('Diamètre non conforme FP2E')
            # Correction proposée : 1er diamètre listé pour la lettre
            if expected_diametres:
                corrections['diametre'] = str(expected_diametres[0])

    except (TypeError, ValueError, IndexError):
        anomalies.append('Erreur de format interne')

    return anomalies, corrections


def check_fp2e_with_suffix(row):
    """
    Vérifie les compteurs avec format FP2E + 1 lettre finale (pour marques autres).
    Ignore la dernière lettre pour l'analyse année/diamètre.
    Renvoie (liste_anomalies, dict_corrections).
    """
    anomalies, corrections = [], {}
    try:
        compteur = str(row['Numéro de compteur']).strip()
        annee_fabrication_val = str(row['Année de fabrication']).strip()
        diametre_val = row['Diametre']

        # Vérifier format FP2E + suffixe
        if not re.match(FP2E_WITH_SUFFIX_REGEX, compteur):
            return [], {}

        # Analyser en ignorant la dernière lettre (suffixe)
        # Millésime = positions 1-2 ; lettre diamètre = position 4
        annee_compteur = compteur[1:3]
        lettre_diam = compteur[4].upper()

        # Année : vide / non numérique / non égale au millésime => propose correction
        if (
            annee_fabrication_val == ''
            or not annee_fabrication_val.isdigit()
            or annee_compteur != annee_fabrication_val.zfill(2)
        ):
            anomalies.append("L'année de millésime n'est pas conforme (FP2E+suffixe)")
            corrections['annee'] = annee_compteur

        # Diamètre attendu depuis la lettre
        expected_diametres = _diam_map_fp2e().get(lettre_diam, [])
        if not isinstance(expected_diametres, list):
            expected_diametres = [expected_diametres]

        if pd.isna(diametre_val) or diametre_val not in expected_diametres:
            anomalies.append("Le diamètre n'est pas conforme (FP2E+suffixe)")
            # Correction proposée : 1er diamètre listé pour la lettre
            if expected_diametres:
                corrections['diametre'] = str(expected_diametres[0])

    except (TypeError, ValueError, IndexError):
        anomalies.append("Le numéro de compteur n'est pas conforme (FP2E+suffixe)")

    return anomalies, corrections


def check_data_radio(df):
    """
    Règles onglet 'Radio' :
      - Marque autorisée (liste blanche : SAPPEL (H)/(C), KAMSTRUP, U Kamstrup),
      - Protocole par marque/année (KAMSTRUP=WMS, SAPPEL<=22=WMS, >22=OMS),
      - Présence champs clés, coordonnées valides,
      - Règles KAMSTRUP (8 car., = tête, numérique, Ø 15-80),
      - Règles SAPPEL (tête DME 15 car., compteur C/H si non manuelle),
      - ITRON (compteur I/D si non manuelle),
      - Déduction Type Compteur (2 lettres issues du n°),
      - Contrôles FP2E détaillés (année/Ø) si condition FP2E remplie,
      - Support FP2E + lettre finale si Traité = 965/455/899.
    Retourne (df_anomalies, compteur_par_type).
    """
    df_with_anomalies = df.copy()

    # Colonnes de corrections (vides par défaut)
    df_with_anomalies['Correction Année'] = ''
    df_with_anomalies['Correction Diamètre'] = ''
    df_with_anomalies['Correction Type Compteur'] = ''
    df_with_anomalies['Correction Marque'] = ''
    df_with_anomalies['Correction Numéro de Tête'] = ''
    df_with_anomalies['Correction Protocole Radio'] = ''

    if 'Type Compteur' not in df_with_anomalies.columns:
        raise ValueError("La colonne 'Type Compteur' est manquante dans votre fichier.")

    # Normalisation année (2 digits, zfill)
    df_with_anomalies['Année de fabrication'] = (
        df_with_anomalies['Année de fabrication']
        .astype(str)
        .replace('nan', '', regex=False)
        .apply(lambda x: str(int(float(x))) if x.replace('.', '', 1).isdigit() and x != '' else x)
        .str.slice(-2)
        .str.zfill(2)
    )

    # Présence colonnes indispensables
    required_columns = [
        'Protocole Radio', 'Marque', 'Numéro de tête', 'Numéro de compteur',
        'Latitude', 'Longitude', 'Commune', 'Année de fabrication', 'Diametre',
        'Mode de relève', 'Type Compteur'
    ]
    if not all(col in df_with_anomalies.columns for col in required_columns):
        missing_columns = [col for col in required_columns if col not in df_with_anomalies.columns]
        raise ValueError(f"Colonnes requises manquantes : {', '.join(missing_columns)}")

    df_with_anomalies['Anomalie'] = ''
    
    # Normalisation Traité (si la colonne existe) pour détecter 965/455/899
    if 'Traité' in df_with_anomalies.columns:
        df_with_anomalies['Traité'] = df_with_anomalies['Traité'].astype(str).replace('nan', '', regex=False)
        is_traite_special = df_with_anomalies['Traité'].str.startswith(('965', '455', '899'), na=False)
    else:
        is_traite_special = pd.Series([False] * len(df_with_anomalies), index=df_with_anomalies.index)

    # Chaînes propres + numerics coord/Ø
    for col in ['Numéro de compteur', 'Numéro de tête', 'Marque', 'Protocole Radio', 'Mode de relève', 'Type Compteur']:
        df_with_anomalies[col] = df_with_anomalies[col].astype(str).replace('nan', '', regex=False)
    df_with_anomalies['Latitude'] = pd.to_numeric(df_with_anomalies['Latitude'], errors='coerce')
    df_with_anomalies['Longitude'] = pd.to_numeric(df_with_anomalies['Longitude'], errors='coerce')

    # Flags marques
    is_kamstrup = df_with_anomalies['Marque'].str.upper() == 'KAMSTRUP'
    is_u_kamstrup = df_with_anomalies['Marque'].str.upper().str.replace(' ', '', regex=False) == 'UKAMSTRUP'
    is_sappel = df_with_anomalies['Marque'].str.upper().isin(['SAPPEL (C)', 'SAPPEL (H)'])
    is_itron = df_with_anomalies['Marque'].str.upper() == 'ITRON'
    annee_fabrication_num = pd.to_numeric(df_with_anomalies['Année de fabrication'], errors='coerce')
    df_with_anomalies['Diametre'] = pd.to_numeric(df_with_anomalies['Diametre'], errors='coerce')
    
    # Détection formats FP2E (standard et avec suffixe)
    has_fp2e_format = df_with_anomalies['Numéro de compteur'].str.match(FP2E_REGEX, na=False)
    has_fp2e_suffix_format = df_with_anomalies['Numéro de compteur'].str.match(FP2E_WITH_SUFFIX_REGEX, na=False)

    # Protocoles attendus
    kamstrup_protocole_incorrect = is_kamstrup & (df_with_anomalies['Protocole Radio'].str.upper() != 'WMS')
    df_with_anomalies.loc[kamstrup_protocole_incorrect, 'Anomalie'] += 'KAMSTRUP: Protocole ≠ WMS / '
    df_with_anomalies.loc[kamstrup_protocole_incorrect, 'Correction Protocole Radio'] = 'WMS'

    sappel_protocole_incorrect_wms = is_sappel & (annee_fabrication_num <= 22) & (df_with_anomalies['Protocole Radio'].str.upper() != 'WMS')
    df_with_anomalies.loc[sappel_protocole_incorrect_wms, 'Anomalie'] += 'SAPPEL: Protocole ≠ WMS (année <= 22) / '
    df_with_anomalies.loc[sappel_protocole_incorrect_wms, 'Correction Protocole Radio'] = 'WMS'

    sappel_protocole_incorrect_oms = is_sappel & (annee_fabrication_num > 22) & (df_with_anomalies['Protocole Radio'].str.upper() != 'OMS')
    df_with_anomalies.loc[sappel_protocole_incorrect_oms, 'Anomalie'] += 'SAPPEL: Protocole ≠ OMS (année > 22) / '
    df_with_anomalies.loc[sappel_protocole_incorrect_oms, 'Correction Protocole Radio'] = 'OMS'

    # Manques / formats GPS
    df_with_anomalies.loc[df_with_anomalies['Marque'].isin(['', 'nan']), 'Anomalie'] += 'Marque manquante / '
    df_with_anomalies.loc[df_with_anomalies['Numéro de compteur'].isin(['', 'nan']), 'Anomalie'] += 'Numéro de compteur manquant / '
    df_with_anomalies.loc[df_with_anomalies['Diametre'].isnull(), 'Anomalie'] += 'Diamètre manquant / '
    df_with_anomalies.loc[annee_fabrication_num.isnull(), 'Anomalie'] += 'Année de fabrication manquante / '

    # Marque autorisée en radiorelève (liste blanche configurable)
    cfg = regles_config.get_config()
    marques_autorisees_radio = cfg.marques_autorisees_norm('Radio')
    marque_normalisee = df_with_anomalies['Marque'].str.upper().str.replace(' ', '', regex=False)
    marque_renseignee = ~df_with_anomalies['Marque'].isin(['', 'nan'])
    df_with_anomalies.loc[
        marque_renseignee & (~marque_normalisee.isin(marques_autorisees_radio)),
        'Anomalie'
    ] += 'Marque non autorisée en radiorelève / '

    # Type Compteur autorisé (liste blanche configurable)
    type_compteur_norm = df_with_anomalies['Type Compteur'].astype(str).str.upper().str.replace(' ', '', regex=False)
    type_compteur_renseigne = ~type_compteur_norm.isin(['', 'NAN'])
    df_with_anomalies.loc[
        type_compteur_renseigne & (~type_compteur_norm.isin(cfg.types_valides_norm())),
        'Anomalie'
    ] += 'Type Compteur non autorisé / '

    tete_manquante = df_with_anomalies['Numéro de tête'].isin(['', 'nan'])

    # Tête exigée (hors cas exclus)
    condition_tete_sappel = (
        tete_manquante
        & (~is_sappel | (annee_fabrication_num >= 22))
        & (df_with_anomalies['Mode de relève'].str.upper() != 'MANUELLE')
        & (~is_kamstrup)
    )
    df_with_anomalies.loc[condition_tete_sappel, 'Anomalie'] += 'Numéro de tête manquant / '

    # KAMSTRUP : si n°compteur = 8 chiffres, tête = compteur
    condition_tete_kamstrup = tete_manquante & is_kamstrup & (df_with_anomalies['Numéro de compteur'].str.match(r'^\d{8}$'))
    df_with_anomalies.loc[condition_tete_kamstrup, 'Anomalie'] += 'Numéro de tête manquant / '
    df_with_anomalies.loc[condition_tete_kamstrup, 'Correction Numéro de Tête'] = df_with_anomalies.loc[condition_tete_kamstrup, 'Numéro de compteur']

    # GPS numériques + bornes plausibles
    df_with_anomalies.loc[
        df_with_anomalies['Latitude'].isnull() | df_with_anomalies['Longitude'].isnull(),
        'Anomalie'
    ] += 'Coordonnées GPS non numériques / '
    df_with_anomalies.loc[
        ((df_with_anomalies['Latitude'] == 0) | (~df_with_anomalies['Latitude'].between(-90, 90)))
        | ((df_with_anomalies['Longitude'] == 0) | (~df_with_anomalies['Longitude'].between(-180, 180))),
        'Anomalie'
    ] += 'Coordonnées GPS invalides / '

    # Règles KAMSTRUP
    # Séparer KAMSTRUP FP2E (commence par U) et KAMSTRUP classique (8 chiffres)
    kamstrup_fp2e = is_kamstrup & df_with_anomalies['Numéro de compteur'].str.startswith('U', na=False)
    kamstrup_classique = is_kamstrup & (~df_with_anomalies['Numéro de compteur'].str.startswith('U', na=False))
    
    # KAMSTRUP Classique (ancien format 8 chiffres)
    kamstrup_valid = kamstrup_classique & (~df_with_anomalies['Numéro de tête'].isin(['', 'nan']))
    df_with_anomalies.loc[kamstrup_classique & (df_with_anomalies['Numéro de compteur'].str.len() != 8), 'Anomalie'] += 'KAMSTRUP: Compteur ≠ 8 caractères / '
    df_with_anomalies.loc[kamstrup_valid & (df_with_anomalies['Numéro de compteur'] != df_with_anomalies['Numéro de tête']), 'Anomalie'] += 'KAMSTRUP: Compteur ≠ Tête / '
    df_with_anomalies.loc[kamstrup_valid & (~df_with_anomalies['Numéro de compteur'].str.isdigit() | ~df_with_anomalies['Numéro de tête'].str.isdigit()), 'Anomalie'] += 'KAMSTRUP: Compteur ou Tête non numérique / '
    _diam_min, _diam_max = cfg.diametre_min_max('KAMSTRUP')
    df_with_anomalies.loc[kamstrup_classique & (~df_with_anomalies['Diametre'].between(_diam_min, _diam_max)), 'Anomalie'] += 'KAMSTRUP: Diamètre hors plage / '
    
    # KAMSTRUP FP2E (nouveau format commençant par U)
    # Vérifier format FP2E (11 caractères ou 12 si Traité spécial)
    kamstrup_fp2e_format_ok = kamstrup_fp2e & (has_fp2e_format | (is_traite_special & has_fp2e_suffix_format))
    kamstrup_fp2e_format_ko = kamstrup_fp2e & (~has_fp2e_format) & (~(is_traite_special & has_fp2e_suffix_format))
    df_with_anomalies.loc[kamstrup_fp2e_format_ko, 'Anomalie'] += 'KAMSTRUP: Format FP2E invalide / '

    # Marque U Kamstrup : compteur conforme FP2E + tête à 8 chiffres
    u_kamstrup_format_ko = is_u_kamstrup & (~has_fp2e_format) & (~(is_traite_special & has_fp2e_suffix_format))
    df_with_anomalies.loc[u_kamstrup_format_ko, 'Anomalie'] += 'U Kamstrup: Format FP2E invalide / '
    u_kamstrup_tete_ko = (
        is_u_kamstrup
        & (~df_with_anomalies['Numéro de tête'].isin(['', 'nan']))
        & (~df_with_anomalies['Numéro de tête'].str.match(r'^\d{8}$'))
    )
    df_with_anomalies.loc[u_kamstrup_tete_ko, 'Anomalie'] += 'U Kamstrup: Tête ≠ 8 chiffres / '

    # Règles SAPPEL / ITRON spécifiques
    df_with_anomalies.loc[
        is_sappel
        & (df_with_anomalies['Numéro de tête'].astype(str).str.upper().str.startswith('DME'))
        & (df_with_anomalies['Numéro de tête'].str.len() != 15),
        'Anomalie'
    ] += 'SAPPEL: Tête DME ≠ 15 caractères / '

    # Longueurs de tête selon Mode/Marque/Type Compteur (configurable)
    appliquer_longueur_tete(df_with_anomalies, 'Radio', cfg)

    df_with_anomalies.loc[
        is_sappel
        & (df_with_anomalies['Mode de relève'].str.upper() != 'MANUELLE')
        & (~df_with_anomalies['Numéro de compteur'].str.startswith(('C', 'H'))),
        'Anomalie'
    ] += 'SAPPEL: Compteur ne commence pas par C ou H / '

    # Marque vs préfixe compteur (cohérences C/H)
    compteur_starts_C = df_with_anomalies['Numéro de compteur'].str.startswith('C')
    marque_not_sappel_C = df_with_anomalies['Marque'].str.upper() != 'SAPPEL (C)'
    df_with_anomalies.loc[is_sappel & compteur_starts_C & marque_not_sappel_C, 'Anomalie'] += 'SAPPEL: Incohérence Marque/Compteur (C) / '
    df_with_anomalies.loc[is_sappel & compteur_starts_C & marque_not_sappel_C, 'Correction Marque'] = 'SAPPEL (C)'

    compteur_starts_H = df_with_anomalies['Numéro de compteur'].str.startswith('H')
    marque_not_sappel_H = df_with_anomalies['Marque'].str.upper() != 'SAPPEL (H)'
    df_with_anomalies.loc[is_sappel & compteur_starts_H & marque_not_sappel_H, 'Anomalie'] += 'SAPPEL: Incohérence Marque/Compteur (H) / '
    df_with_anomalies.loc[is_sappel & compteur_starts_H & marque_not_sappel_H, 'Correction Marque'] = 'SAPPEL (H)'

    df_with_anomalies.loc[
        is_itron
        & (df_with_anomalies['Mode de relève'].str.upper() != 'MANUELLE')
        & (~df_with_anomalies['Numéro de compteur'].str.startswith(('I', 'D'))),
        'Anomalie'
    ] += 'ITRON: Compteur ne commence pas par I ou D / '

    # Déduction du 'Type Compteur' attendu via 1er et 4e char du n°
    is_brand_ok = is_sappel | is_itron
    is_len_ok = df_with_anomalies['Numéro de compteur'].str.len() == 11
    starts_with_letter = df_with_anomalies['Numéro de compteur'].str[0].str.isalpha()
    fourth_is_letter = df_with_anomalies['Numéro de compteur'].str[3].str.isalpha()
    condition_type_compteur = is_brand_ok & is_len_ok & starts_with_letter & fourth_is_letter

    rows_to_check = df_with_anomalies[condition_type_compteur].copy()
    if not rows_to_check.empty:
        # SAPPEL : type = (c0 + c3)
        sappel_rows = rows_to_check[rows_to_check['Marque'].str.upper().isin(['SAPPEL (C)', 'SAPPEL (H)'])]
        if not sappel_rows.empty:
            correct_type_sappel = sappel_rows['Numéro de compteur'].str[0] + sappel_rows['Numéro de compteur'].str[3]
            incorrect_mask_sappel = sappel_rows['Type Compteur'] != correct_type_sappel
            incorrect_indices_sappel = sappel_rows[incorrect_mask_sappel].index
            if not incorrect_indices_sappel.empty:
                df_with_anomalies.loc[incorrect_indices_sappel, 'Anomalie'] += 'Incohérence Type Compteur / '
                df_with_anomalies.loc[incorrect_indices_sappel, 'Correction Type Compteur'] = correct_type_sappel[incorrect_mask_sappel]

        # ITRON : type = 'I' + c3
        itron_rows = rows_to_check[rows_to_check['Marque'].str.upper() == 'ITRON']
        if not itron_rows.empty:
            correct_type_itron = 'I' + itron_rows['Numéro de compteur'].str[3]
            incorrect_mask_itron = itron_rows['Type Compteur'] != correct_type_itron
            incorrect_indices_itron = itron_rows[incorrect_mask_itron].index
            if not incorrect_indices_itron.empty:
                df_with_anomalies.loc[incorrect_indices_itron, 'Anomalie'] += 'Incohérence Type Compteur / '
                df_with_anomalies.loc[incorrect_indices_itron, 'Correction Type Compteur'] = correct_type_itron[incorrect_mask_itron]

    # Détermine où appliquer la vérif FP2E détaillée
    # Radio non-manuelle SAPPEL / Manuel si format FP2E / KAMSTRUP FP2E
    sappel_non_manuelle_fp2e = is_sappel & (df_with_anomalies['Mode de relève'].str.upper() != 'MANUELLE')
    manuelle_format_ok = (df_with_anomalies['Mode de relève'].str.upper() == 'MANUELLE') & (has_fp2e_format | (is_traite_special & has_fp2e_suffix_format))
    kamstrup_fp2e_check = kamstrup_fp2e & has_fp2e_format
    u_kamstrup_fp2e_check = is_u_kamstrup & has_fp2e_format
    fp2e_check_condition = sappel_non_manuelle_fp2e | manuelle_format_ok | kamstrup_fp2e_check | u_kamstrup_fp2e_check

    fp2e_results = df_with_anomalies[fp2e_check_condition & has_fp2e_format].apply(check_fp2e_details_radio, axis=1)
    for index, result in fp2e_results.items():
        anomalies, corrections = result
        if anomalies:
            df_with_anomalies.loc[index, 'Anomalie'] += ' / '.join(anomalies) + ' / '
        if 'annee' in corrections:
            df_with_anomalies.loc[index, 'Correction Année'] = corrections['annee']
        if 'diametre' in corrections:
            df_with_anomalies.loc[index, 'Correction Diamètre'] = corrections['diametre']
    
    # Contrôles FP2E avec suffixe si Traité = 965/455/899 (y compris KAMSTRUP)
    kamstrup_fp2e_suffix_check = kamstrup_fp2e & has_fp2e_suffix_format & is_traite_special
    u_kamstrup_fp2e_suffix_check = is_u_kamstrup & has_fp2e_suffix_format & is_traite_special
    fp2e_suffix_check = (sappel_non_manuelle_fp2e | manuelle_format_ok | kamstrup_fp2e_suffix_check | u_kamstrup_fp2e_suffix_check) & has_fp2e_suffix_format & is_traite_special
    fp2e_suffix_results = df_with_anomalies[fp2e_suffix_check].apply(check_fp2e_with_suffix, axis=1)
    for index, result in fp2e_suffix_results.items():
        anomalies, corrections = result
        if anomalies:
            df_with_anomalies.loc[index, 'Anomalie'] += ' / '.join(anomalies) + ' / '
        if 'annee' in corrections:
            df_with_anomalies.loc[index, 'Correction Année'] = corrections['annee']
        if 'diametre' in corrections:
            df_with_anomalies.loc[index, 'Correction Diamètre'] = corrections['diametre']

    # Nettoyage du séparateur
    df_with_anomalies['Anomalie'] = df_with_anomalies['Anomalie'].str.strip().str.rstrip(' /')

    # Extraction des lignes en anomalie (ou corrections proposées)
    anomalies_df = df_with_anomalies[
        (df_with_anomalies['Anomalie'] != '')
        | (df_with_anomalies['Correction Année'] != '')
        | (df_with_anomalies['Correction Diamètre'] != '')
        | (df_with_anomalies['Correction Type Compteur'] != '')
        | (df_with_anomalies['Correction Marque'] != '')
        | (df_with_anomalies['Correction Numéro de Tête'] != '')
        | (df_with_anomalies['Correction Protocole Radio'] != '')
    ].copy()

    anomalies_df.reset_index(inplace=True)
    anomalies_df.rename(columns={'index': 'Index original'}, inplace=True)

    # Repositionne les colonnes 'Correction *' près de leur champ d'origine
    try:
        cols = list(anomalies_df.columns)
        for c in ['Correction Année', 'Correction Diamètre', 'Correction Type Compteur',
                  'Correction Marque', 'Correction Numéro de Tête', 'Correction Protocole Radio']:
            cols.remove(c)
        pos_annee = cols.index('Année de fabrication') + 1; cols.insert(pos_annee, 'Correction Année')
        pos_diametre = cols.index('Diametre') + 1; cols.insert(pos_diametre, 'Correction Diamètre')
        pos_type = cols.index('Type Compteur') + 1; cols.insert(pos_type, 'Correction Type Compteur')
        pos_marque = cols.index('Marque') + 1; cols.insert(pos_marque, 'Correction Marque')
        pos_tete = cols.index('Numéro de tête') + 1; cols.insert(pos_tete, 'Correction Numéro de Tête')
        pos_protocole = cols.index('Protocole Radio') + 1; cols.insert(pos_protocole, 'Correction Protocole Radio')
        anomalies_df = anomalies_df[cols]
    except ValueError:
        pass

    # Compteur par type d'anomalie (explosion du libellé)
    return anomalies_df, anomalies_df['Anomalie'].str.split(' / ').explode().value_counts()


def check_data_tele(df):
    """
    Règles onglet 'Télé' :
      - Marque autorisée (liste blanche : INTEGRA, ITRON, KAIFA, KAMSTRUP,
        SAPPEL (C)/(H), SENSUS, SOCAM, U Kamstrup),
      - Protocole selon préfixe Traité (312/455/863/895/903/956 => LRA, sinon SGX) si non manuelle,
      - Présence champs clés, coordonnées valides,
      - KAMSTRUP / SAPPEL / ITRON : longueurs de tête, cohérences,
      - Déduction Type Compteur,
      - Contrôles FP2E détaillés,
      - Rappels manuels ITRON/SAPPEL sur préfixes I/D et C/H,
      - Support FP2E + lettre finale si Traité = 965/455/899.
    """
    df_with_anomalies = df.copy()

    df_with_anomalies['Correction Année'] = ''
    df_with_anomalies['Correction Diamètre'] = ''
    df_with_anomalies['Correction Type Compteur'] = ''
    df_with_anomalies['Correction Marque'] = ''
    df_with_anomalies['Correction Numéro de Tête'] = ''
    df_with_anomalies['Correction Protocole Radio'] = ''

    if 'Type Compteur' not in df_with_anomalies.columns:
        raise ValueError("La colonne 'Type Compteur' est manquante dans votre fichier.")

    # Normalisation année
    df_with_anomalies['Année de fabrication'] = (
        df_with_anomalies['Année de fabrication']
        .astype(str).replace('nan', '', regex=False)
        .apply(lambda x: str(int(float(x))) if x.replace('.', '', 1).isdigit() and x != '' else x)
        .str.slice(-2).str.zfill(2)
    )

    required_columns = [
        'Protocole Radio', 'Marque', 'Numéro de compteur', 'Numéro de tête',
        'Latitude', 'Longitude', 'Année de fabrication', 'Diametre', 'Traité',
        'Mode de relève', 'Type Compteur'
    ]
    if not all(col in df_with_anomalies.columns for col in required_columns):
        missing = [col for col in required_columns if col not in df_with_anomalies.columns]
        raise ValueError(f"Colonnes requises manquantes : {', '.join(missing)}")

    df_with_anomalies['Anomalie'] = ''

    # Normalisations
    for col in ['Numéro de compteur', 'Numéro de tête', 'Marque', 'Protocole Radio', 'Traité', 'Mode de relève', 'Type Compteur']:
        df_with_anomalies[col] = df_with_anomalies[col].astype(str).replace('nan', '', regex=False)
    df_with_anomalies['Latitude'] = pd.to_numeric(df_with_anomalies['Latitude'], errors='coerce')
    df_with_anomalies['Longitude'] = pd.to_numeric(df_with_anomalies['Longitude'], errors='coerce')
    df_with_anomalies['Diametre'] = pd.to_numeric(df_with_anomalies['Diametre'], errors='coerce')

    # Flags
    is_kamstrup = df_with_anomalies['Marque'].str.upper() == 'KAMSTRUP'
    is_u_kamstrup = df_with_anomalies['Marque'].str.upper().str.replace(' ', '', regex=False) == 'UKAMSTRUP'
    is_sappel = df_with_anomalies['Marque'].str.upper().isin(['SAPPEL (C)', 'SAPPEL (H)', 'SAPPEL(C)'])
    is_itron = df_with_anomalies['Marque'].str.upper() == 'ITRON'
    is_kaifa = df_with_anomalies['Marque'].str.upper() == 'KAIFA'
    is_mode_manuelle = df_with_anomalies['Mode de relève'].str.upper() == 'MANUELLE'
    annee_fabrication_num = pd.to_numeric(df_with_anomalies['Année de fabrication'], errors='coerce')
    
    # Détection Traité spécial (965/455/899)
    is_traite_special = df_with_anomalies['Traité'].str.startswith(('965', '455', '899'), na=False)
    
    # Détection formats FP2E (standard et avec suffixe)
    has_fp2e_format = df_with_anomalies['Numéro de compteur'].str.match(FP2E_REGEX, na=False)
    has_fp2e_suffix_format = df_with_anomalies['Numéro de compteur'].str.match(FP2E_WITH_SUFFIX_REGEX, na=False)

    # Protocole attendu (non manuelle) — préfixes LRA configurables ; le reste en SGX
    cfg = regles_config.get_config()
    traite_lra_condition = df_with_anomalies['Traité'].str.startswith(cfg.traites_lra_tuple(), na=False)
    protocole_incorrect_lra = (~is_mode_manuelle) & traite_lra_condition & (df_with_anomalies['Protocole Radio'].str.upper() != 'LRA')
    df_with_anomalies.loc[protocole_incorrect_lra, 'Anomalie'] += 'Protocole incorrect (devrait être LRA) / '
    df_with_anomalies.loc[protocole_incorrect_lra, 'Correction Protocole Radio'] = 'LRA'

    protocole_incorrect_sgx = (~is_mode_manuelle) & (~traite_lra_condition) & (df_with_anomalies['Protocole Radio'].str.upper() != 'SGX')
    df_with_anomalies.loc[protocole_incorrect_sgx, 'Anomalie'] += 'Protocole incorrect (devrait être SGX) / '
    df_with_anomalies.loc[protocole_incorrect_sgx, 'Correction Protocole Radio'] = 'SGX'

    # Manques / GPS
    df_with_anomalies.loc[df_with_anomalies['Marque'].isin(['', 'nan']), 'Anomalie'] += 'Marque manquante / '
    df_with_anomalies.loc[df_with_anomalies['Numéro de compteur'].isin(['', 'nan']), 'Anomalie'] += 'Numéro de compteur manquant / '
    df_with_anomalies.loc[df_with_anomalies['Diametre'].isnull(), 'Anomalie'] += 'Diamètre manquant / '
    df_with_anomalies.loc[annee_fabrication_num.isnull(), 'Anomalie'] += 'Année de fabrication manquante / '

    # Marque autorisée en télérelève (liste blanche configurable)
    marques_autorisees_tele = cfg.marques_autorisees_norm('Tele')
    marque_normalisee = df_with_anomalies['Marque'].str.upper().str.replace(' ', '', regex=False)
    marque_renseignee = ~df_with_anomalies['Marque'].isin(['', 'nan'])
    df_with_anomalies.loc[
        marque_renseignee & (~marque_normalisee.isin(marques_autorisees_tele)),
        'Anomalie'
    ] += 'Marque non autorisée en télérelève / '

    # Type Compteur autorisé (liste blanche configurable)
    type_compteur_norm = df_with_anomalies['Type Compteur'].astype(str).str.upper().str.replace(' ', '', regex=False)
    type_compteur_renseigne = ~type_compteur_norm.isin(['', 'NAN'])
    df_with_anomalies.loc[
        type_compteur_renseigne & (~type_compteur_norm.isin(cfg.types_valides_norm())),
        'Anomalie'
    ] += 'Type Compteur non autorisé / '

    # Tête requise (hors KAMSTRUP/KAIFA et hors manuelle)
    df_with_anomalies.loc[
        df_with_anomalies['Numéro de tête'].isin(['', 'nan']) & (~is_kamstrup) & (~is_kaifa) & (~is_mode_manuelle),
        'Anomalie'
    ] += 'Numéro de tête manquant / '

    # GPS validations
    df_with_anomalies.loc[
        df_with_anomalies['Latitude'].isnull() | df_with_anomalies['Longitude'].isnull(),
        'Anomalie'
    ] += 'Coordonnées GPS non numériques / '
    df_with_anomalies.loc[
        ((df_with_anomalies['Latitude'] == 0) | (~df_with_anomalies['Latitude'].between(-90, 90)))
        | ((df_with_anomalies['Longitude'] == 0) | (~df_with_anomalies['Longitude'].between(-180, 180))),
        'Anomalie'
    ] += 'Coordonnées GPS invalides / '

    # KAMSTRUP
    # Séparer KAMSTRUP FP2E (commence par U) et KAMSTRUP classique (8 chiffres)
    kamstrup_fp2e = is_kamstrup & df_with_anomalies['Numéro de compteur'].str.startswith('U', na=False)
    kamstrup_classique = is_kamstrup & (~df_with_anomalies['Numéro de compteur'].str.startswith('U', na=False))
    
    # KAMSTRUP Classique (ancien format 8 chiffres)
    kamstrup_valid = kamstrup_classique & (~df_with_anomalies['Numéro de tête'].isin(['', 'nan']))
    df_with_anomalies.loc[kamstrup_classique & (df_with_anomalies['Numéro de compteur'].str.len() != 8), 'Anomalie'] += 'KAMSTRUP: Compteur ≠ 8 caractères / '
    df_with_anomalies.loc[kamstrup_valid & (df_with_anomalies['Numéro de compteur'] != df_with_anomalies['Numéro de tête']), 'Anomalie'] += 'KAMSTRUP: Compteur ≠ Tête / '
    df_with_anomalies.loc[kamstrup_valid & (~df_with_anomalies['Numéro de compteur'].str.isdigit() | ~df_with_anomalies['Numéro de tête'].str.isdigit()), 'Anomalie'] += 'KAMSTRUP: Compteur ou Tête non numérique / '
    _diam_min, _diam_max = cfg.diametre_min_max('KAMSTRUP')
    df_with_anomalies.loc[kamstrup_classique & (~df_with_anomalies['Diametre'].between(_diam_min, _diam_max)), 'Anomalie'] += f'KAMSTRUP: Diamètre hors de la plage [{_diam_min}, {_diam_max}] / '
    
    # KAMSTRUP FP2E (nouveau format commençant par U)
    # Vérifier format FP2E (11 caractères ou 12 si Traité spécial)
    kamstrup_fp2e_format_ok = kamstrup_fp2e & (has_fp2e_format | (is_traite_special & has_fp2e_suffix_format))
    kamstrup_fp2e_format_ko = kamstrup_fp2e & (~has_fp2e_format) & (~(is_traite_special & has_fp2e_suffix_format))
    df_with_anomalies.loc[kamstrup_fp2e_format_ko, 'Anomalie'] += 'KAMSTRUP: Format FP2E invalide / '

    # Marque U Kamstrup : compteur conforme FP2E + tête à 8 chiffres
    u_kamstrup_format_ko = is_u_kamstrup & (~has_fp2e_format) & (~(is_traite_special & has_fp2e_suffix_format))
    df_with_anomalies.loc[u_kamstrup_format_ko, 'Anomalie'] += 'U Kamstrup: Format FP2E invalide / '
    u_kamstrup_tete_ko = (
        is_u_kamstrup
        & (~df_with_anomalies['Numéro de tête'].isin(['', 'nan']))
        & (~df_with_anomalies['Numéro de tête'].str.match(r'^\d{8}$'))
    )
    df_with_anomalies.loc[u_kamstrup_tete_ko, 'Anomalie'] += 'U Kamstrup: Tête ≠ 8 chiffres / '

    # Longueurs de tête selon Mode/Marque/Type Compteur (configurable)
    # Couvre SAPPEL (16, ou 15 pour SEN3) et ITRON (8).
    appliquer_longueur_tete(df_with_anomalies, 'Tele', cfg)

    # Cohérences marque vs préfixe
    compteur_starts_C = df_with_anomalies['Numéro de compteur'].str.startswith('C')
    marque_not_sappel_C = df_with_anomalies['Marque'].str.upper() != 'SAPPEL (C)'
    df_with_anomalies.loc[is_sappel & compteur_starts_C & marque_not_sappel_C, 'Anomalie'] += 'SAPPEL: Incohérence Marque/Compteur (C) / '
    df_with_anomalies.loc[is_sappel & compteur_starts_C & marque_not_sappel_C, 'Correction Marque'] = 'SAPPEL (C)'

    compteur_starts_H = df_with_anomalies['Numéro de compteur'].str.startswith('H')
    marque_not_sappel_H = df_with_anomalies['Marque'].str.upper() != 'SAPPEL (H)'
    df_with_anomalies.loc[is_sappel & compteur_starts_H & marque_not_sappel_H, 'Anomalie'] += 'SAPPEL: Incohérence Marque/Compteur (H) / '
    df_with_anomalies.loc[is_sappel & compteur_starts_H & marque_not_sappel_H, 'Correction Marque'] = 'SAPPEL (H)'

    # Déduction 'Type Compteur'
    is_brand_ok = is_sappel | is_itron
    is_len_ok = df_with_anomalies['Numéro de compteur'].str.len() == 11
    starts_with_letter = df_with_anomalies['Numéro de compteur'].str[0].str.isalpha()
    fourth_is_letter = df_with_anomalies['Numéro de compteur'].str[3].str.isalpha()
    condition_type_compteur = is_brand_ok & is_len_ok & starts_with_letter & fourth_is_letter

    rows_to_check = df_with_anomalies[condition_type_compteur].copy()
    if not rows_to_check.empty:
        sappel_rows = rows_to_check[rows_to_check['Marque'].str.upper().isin(['SAPPEL (C)', 'SAPPEL (H)', 'SAPPEL(C)'])]
        if not sappel_rows.empty:
            correct_type_sappel = sappel_rows['Numéro de compteur'].str[0] + sappel_rows['Numéro de compteur'].str[3]
            incorrect_mask_sappel = sappel_rows['Type Compteur'] != correct_type_sappel
            incorrect_indices_sappel = sappel_rows[incorrect_mask_sappel].index
            if not incorrect_indices_sappel.empty:
                df_with_anomalies.loc[incorrect_indices_sappel, 'Anomalie'] += 'Incohérence Type Compteur / '
                df_with_anomalies.loc[incorrect_indices_sappel, 'Correction Type Compteur'] = correct_type_sappel[incorrect_mask_sappel]

        itron_rows = rows_to_check[rows_to_check['Marque'].str.upper() == 'ITRON']
        if not itron_rows.empty:
            correct_type_itron = 'I' + itron_rows['Numéro de compteur'].str[3]
            incorrect_mask_itron = itron_rows['Type Compteur'] != correct_type_itron
            incorrect_indices_itron = itron_rows[incorrect_mask_itron].index
            if not incorrect_indices_itron.empty:
                df_with_anomalies.loc[incorrect_indices_itron, 'Anomalie'] += 'Incohérence Type Compteur / '
                df_with_anomalies.loc[incorrect_indices_itron, 'Correction Type Compteur'] = correct_type_itron[incorrect_mask_itron]

    # Contrôles FP2E (zone télé) - Standard (y compris KAMSTRUP FP2E)
    kamstrup_fp2e_check = kamstrup_fp2e & has_fp2e_format
    u_kamstrup_fp2e_check = is_u_kamstrup & has_fp2e_format
    fp2e_condition = ((is_sappel | is_itron) & (~is_mode_manuelle)) | (is_mode_manuelle & has_fp2e_format) | kamstrup_fp2e_check | u_kamstrup_fp2e_check
    fp2e_results = df_with_anomalies[fp2e_condition & has_fp2e_format].apply(check_fp2e_details_tele, axis=1)

    for index, result in fp2e_results.items():
        anomalies, corrections = result
        if anomalies:
            df_with_anomalies.loc[index, 'Anomalie'] += ' / '.join(anomalies) + ' / '
        if 'annee' in corrections:
            df_with_anomalies.loc[index, 'Correction Année'] = corrections['annee']
        if 'diametre' in corrections:
            df_with_anomalies.loc[index, 'Correction Diamètre'] = corrections['diametre']
    
    # Contrôles FP2E avec suffixe si Traité = 965/455/899 (y compris KAMSTRUP)
    kamstrup_fp2e_suffix_check = kamstrup_fp2e & has_fp2e_suffix_format & is_traite_special
    u_kamstrup_fp2e_suffix_check = is_u_kamstrup & has_fp2e_suffix_format & is_traite_special
    fp2e_suffix_condition = ((((is_sappel | is_itron) & (~is_mode_manuelle)) | is_mode_manuelle) | kamstrup_fp2e_suffix_check | u_kamstrup_fp2e_suffix_check) & has_fp2e_suffix_format & is_traite_special
    fp2e_suffix_results = df_with_anomalies[fp2e_suffix_condition].apply(check_fp2e_with_suffix, axis=1)
    
    for index, result in fp2e_suffix_results.items():
        anomalies, corrections = result
        if anomalies:
            df_with_anomalies.loc[index, 'Anomalie'] += ' / '.join(anomalies) + ' / '
        if 'annee' in corrections:
            df_with_anomalies.loc[index, 'Correction Année'] = corrections['annee']
        if 'diametre' in corrections:
            df_with_anomalies.loc[index, 'Correction Diamètre'] = corrections['diametre']

    # Rappels préfixes en mode manuel (si FP2E standard uniquement)
    is_fp2e_compliant = has_fp2e_format
    df_with_anomalies.loc[
        is_mode_manuelle & is_itron & is_fp2e_compliant & (~df_with_anomalies['Numéro de compteur'].str.lower().str.startswith(('i', 'd'), na=False)),
        'Anomalie'
    ] += 'ITRON manuel: doit commencer par "I" ou "D" / '
    df_with_anomalies.loc[
        is_mode_manuelle & is_sappel & is_fp2e_compliant & (~df_with_anomalies['Numéro de compteur'].str.lower().str.startswith(('c', 'h'), na=False)),
        'Anomalie'
    ] += 'SAPPEL manuel: doit commencer par "C" ou "H" / '

    # Sortie anomalies + mise en forme colonnes 'Correction *'
    df_with_anomalies['Anomalie'] = df_with_anomalies['Anomalie'].str.strip().str.rstrip(' /')

    anomalies_df = df_with_anomalies[
        (df_with_anomalies['Anomalie'] != '')
        | (df_with_anomalies['Correction Année'] != '')
        | (df_with_anomalies['Correction Diamètre'] != '')
        | (df_with_anomalies['Correction Type Compteur'] != '')
        | (df_with_anomalies['Correction Marque'] != '')
        | (df_with_anomalies['Correction Numéro de Tête'] != '')
        | (df_with_anomalies['Correction Protocole Radio'] != '')
    ].copy()
    anomalies_df.reset_index(inplace=True)
    anomalies_df.rename(columns={'index': 'Index original'}, inplace=True)

    try:
        cols = list(anomalies_df.columns)
        for c in ['Correction Année', 'Correction Diamètre', 'Correction Type Compteur',
                  'Correction Marque', 'Correction Numéro de Tête', 'Correction Protocole Radio']:
            cols.remove(c)
        pos_annee = cols.index('Année de fabrication') + 1; cols.insert(pos_annee, 'Correction Année')
        pos_diametre = cols.index('Diametre') + 1; cols.insert(pos_diametre, 'Correction Diamètre')
        pos_type = cols.index('Type Compteur') + 1; cols.insert(pos_type, 'Correction Type Compteur')
        pos_marque = cols.index('Marque') + 1; cols.insert(pos_marque, 'Correction Marque')
        pos_tete = cols.index('Numéro de tête') + 1; cols.insert(pos_tete, 'Correction Numéro de Tête')
        pos_protocole = cols.index('Protocole Radio') + 1; cols.insert(pos_protocole, 'Correction Protocole Radio')
        anomalies_df = anomalies_df[cols]
    except ValueError:
        pass

    return anomalies_df, anomalies_df['Anomalie'].str.split(' / ').explode().value_counts()


def check_data_manuelle(df):
    """
    Règles onglet 'Manuelle' :
      - Coordonnées numériques/valides,
      - Cohérence FP2E pour SAPPEL/ITRON,
      - Cohérence Marque vs préfixes C/H/I/D,
      - Déduction Type Compteur,
      - Contrôles FP2E détaillés (année/Ø),
      - NOUVEAU: Support FP2E + lettre finale si Traité commence par 965.
    """
    required_cols = ['Latitude', 'Longitude', 'Numéro de compteur', 'Marque', 'Année de fabrication', 'Diametre', 'Type Compteur']
    if not all(col in df.columns for col in required_cols):
        missing = [col for col in required_cols if col not in df.columns]
        raise ValueError(f"Colonnes requises manquantes : {', '.join(missing)}")

    df_with_anomalies = df.copy()
    df_with_anomalies['Anomalie'] = ''
    df_with_anomalies['Correction Année'] = ''
    df_with_anomalies['Correction Diamètre'] = ''
    df_with_anomalies['Correction Marque'] = ''
    df_with_anomalies['Correction Type Compteur'] = ''

    # Normalisations
    df_with_anomalies['Année de fabrication'] = (
        df_with_anomalies['Année de fabrication'].astype(str)
        .replace('nan', '', regex=False)
        .apply(lambda x: str(int(float(x))) if x.replace('.', '', 1).isdigit() and x != '' else x)
        .str.slice(-2).str.zfill(2)
    )
    df_with_anomalies['Latitude'] = pd.to_numeric(df_with_anomalies['Latitude'], errors='coerce')
    df_with_anomalies['Longitude'] = pd.to_numeric(df_with_anomalies['Longitude'], errors='coerce')
    df_with_anomalies['Diametre'] = pd.to_numeric(df_with_anomalies['Diametre'], errors='coerce')

    # Normalisation Traité (si la colonne existe)
    if 'Traité' in df_with_anomalies.columns:
        df_with_anomalies['Traité'] = df_with_anomalies['Traité'].astype(str).replace('nan', '', regex=False)
        is_traite_special = df_with_anomalies['Traité'].str.startswith(('965', '455', '899'), na=False)
    else:
        is_traite_special = pd.Series([False] * len(df_with_anomalies), index=df_with_anomalies.index)

    # GPS
    df_with_anomalies.loc[df_with_anomalies['Latitude'].isnull() | df_with_anomalies['Longitude'].isnull(), 'Anomalie'] += 'Coordonnées GPS non numériques / '
    coord_invalid = ((df_with_anomalies['Latitude'] == 0) | (~df_with_anomalies['Latitude'].between(-90, 90))) | ((df_with_anomalies['Longitude'] == 0) | (~df_with_anomalies['Longitude'].between(-180, 180)))
    df_with_anomalies.loc[coord_invalid, 'Anomalie'] += 'Coordonnées GPS invalides / '

    # Type Compteur autorisé (liste blanche configurable)
    cfg = regles_config.get_config()
    type_compteur_norm = df_with_anomalies['Type Compteur'].astype(str).str.upper().str.replace(' ', '', regex=False)
    type_compteur_renseigne = ~type_compteur_norm.isin(['', 'NAN'])
    df_with_anomalies.loc[
        type_compteur_renseigne & (~type_compteur_norm.isin(cfg.types_valides_norm())),
        'Anomalie'
    ] += 'Type Compteur non autorisé / '

    # Flags marques
    is_sappel = df_with_anomalies['Marque'].str.upper().isin(['SAPPEL (C)', 'SAPPEL (H)'])
    is_itron = df_with_anomalies['Marque'].str.upper() == 'ITRON'
    
    # Format compteur
    has_fp2e_format = df_with_anomalies['Numéro de compteur'].str.match(FP2E_REGEX, na=False)
    has_fp2e_suffix_format = df_with_anomalies['Numéro de compteur'].str.match(FP2E_WITH_SUFFIX_REGEX, na=False)

    # FP2E attendu pour SAPPEL/ITRON
    # Si Traité = 965/455/899 et format FP2E+suffixe, c'est OK, sinon on exige FP2E standard
    condition_fp2e_ok = has_fp2e_format | (is_traite_special & has_fp2e_suffix_format)
    df_with_anomalies.loc[(is_sappel | is_itron) & (~condition_fp2e_ok), 'Anomalie'] += 'Compteur non-FP2E pour SAPPEL/ITRON / '

    # Cohérences Marque vs préfixes (C/H/I/D) - format FP2E standard
    compteur_starts_C = df_with_anomalies['Numéro de compteur'].str.startswith('C')
    marque_not_sappel_C = df_with_anomalies['Marque'].str.upper() != 'SAPPEL (C)'
    df_with_anomalies.loc[has_fp2e_format & compteur_starts_C & marque_not_sappel_C, 'Anomalie'] += 'SAPPEL: Incohérence Marque/Compteur (C) / '
    df_with_anomalies.loc[has_fp2e_format & compteur_starts_C & marque_not_sappel_C, 'Correction Marque'] = 'SAPPEL (C)'

    compteur_starts_H = df_with_anomalies['Numéro de compteur'].str.startswith('H')
    marque_not_sappel_H = df_with_anomalies['Marque'].str.upper() != 'SAPPEL (H)'
    df_with_anomalies.loc[has_fp2e_format & compteur_starts_H & marque_not_sappel_H, 'Anomalie'] += 'SAPPEL: Incohérence Marque/Compteur (H) / '
    df_with_anomalies.loc[has_fp2e_format & compteur_starts_H & marque_not_sappel_H, 'Correction Marque'] = 'SAPPEL (H)'

    compteur_starts_ID = df_with_anomalies['Numéro de compteur'].str.startswith(('I', 'D'))
    marque_not_itron = df_with_anomalies['Marque'].str.upper() != 'ITRON'
    df_with_anomalies.loc[has_fp2e_format & compteur_starts_ID & marque_not_itron, 'Anomalie'] += 'ITRON: Incohérence Marque/Compteur / '
    df_with_anomalies.loc[has_fp2e_format & compteur_starts_ID & marque_not_itron, 'Correction Marque'] = 'ITRON'

    # Détails FP2E standard (année/Ø)
    fp2e_results = df_with_anomalies[has_fp2e_format].apply(check_fp2e_details_radio, axis=1)
    for index, result in fp2e_results.items():
        anomalies, corrections = result
        if anomalies:
            df_with_anomalies.loc[index, 'Anomalie'] += ' / '.join(anomalies) + ' / '
        if 'annee' in corrections:
            df_with_anomalies.loc[index, 'Correction Année'] = corrections['annee']
        if 'diametre' in corrections:
            df_with_anomalies.loc[index, 'Correction Diamètre'] = corrections['diametre']

    # NOUVEAU: Détails FP2E avec suffixe si Traité = 965/455/899
    fp2e_suffix_condition = has_fp2e_suffix_format & is_traite_special
    
    fp2e_suffix_results = df_with_anomalies[fp2e_suffix_condition].apply(check_fp2e_with_suffix, axis=1)
    for index, result in fp2e_suffix_results.items():
        anomalies, corrections = result
        if anomalies:
            df_with_anomalies.loc[index, 'Anomalie'] += ' / '.join(anomalies) + ' / '
        if 'annee' in corrections:
            df_with_anomalies.loc[index, 'Correction Année'] = corrections['annee']
        if 'diametre' in corrections:
            df_with_anomalies.loc[index, 'Correction Diamètre'] = corrections['diametre']

    # Déduction 'Type Compteur' (SAPPEL : c0+c3 ; ITRON : I+c3)
    starts_with_key_letter = df_with_anomalies['Numéro de compteur'].str.startswith(('C', 'H', 'I', 'D'))
    condition_type_compteur = has_fp2e_format & starts_with_key_letter
    rows_to_check = df_with_anomalies[condition_type_compteur].copy()

    if not rows_to_check.empty:
        sappel_mask = rows_to_check['Numéro de compteur'].str.startswith(('C', 'H'))
        sappel_rows = rows_to_check[sappel_mask]
        if not sappel_rows.empty:
            correct_type_sappel = sappel_rows['Numéro de compteur'].str[0] + sappel_rows['Numéro de compteur'].str[3]
            incorrect_mask_sappel = sappel_rows['Type Compteur'] != correct_type_sappel
            incorrect_indices_sappel = sappel_rows[incorrect_mask_sappel].index
            if not incorrect_indices_sappel.empty:
                df_with_anomalies.loc[incorrect_indices_sappel, 'Anomalie'] += 'Incohérence Type Compteur / '
                df_with_anomalies.loc[incorrect_indices_sappel, 'Correction Type Compteur'] = correct_type_sappel[incorrect_mask_sappel]

        itron_mask = rows_to_check['Numéro de compteur'].str.startswith(('I', 'D'))
        itron_rows = rows_to_check[itron_mask]
        if not itron_rows.empty:
            correct_type_itron = 'I' + itron_rows['Numéro de compteur'].str[3]
            incorrect_mask_itron = itron_rows['Type Compteur'] != correct_type_itron
            incorrect_indices_itron = itron_rows[incorrect_mask_itron].index
            if not incorrect_indices_itron.empty:
                df_with_anomalies.loc[incorrect_indices_itron, 'Anomalie'] += 'Incohérence Type Compteur / '
                df_with_anomalies.loc[incorrect_indices_itron, 'Correction Type Compteur'] = correct_type_itron[incorrect_mask_itron]

    # Sortie anomalies
    df_with_anomalies['Anomalie'] = df_with_anomalies['Anomalie'].str.strip().str.rstrip(' /')
    anomalies_df = df_with_anomalies[
        (df_with_anomalies['Anomalie'] != '')
        | (df_with_anomalies['Correction Année'] != '')
        | (df_with_anomalies['Correction Diamètre'] != '')
        | (df_with_anomalies['Correction Marque'] != '')
        | (df_with_anomalies['Correction Type Compteur'] != '')
    ].copy()

    if not anomalies_df.empty:
        anomalies_df.reset_index(inplace=True)
        anomalies_df.rename(columns={'index': 'Index original'}, inplace=True)
        try:
            cols = list(anomalies_df.columns)
            for c in ['Correction Année', 'Correction Diamètre', 'Correction Marque', 'Correction Type Compteur']:
                cols.remove(c)
            pos_annee = cols.index('Année de fabrication') + 1; cols.insert(pos_annee, 'Correction Année')
            pos_diametre = cols.index('Diametre') + 1; cols.insert(pos_diametre, 'Correction Diamètre')
            pos_marque = cols.index('Marque') + 1; cols.insert(pos_marque, 'Correction Marque')
            pos_type = cols.index('Type Compteur') + 1; cols.insert(pos_type, 'Correction Type Compteur')
            anomalies_df = anomalies_df[cols]
        except ValueError:
            pass

    anomaly_counter = anomalies_df['Anomalie'].str.split(' / ').explode().value_counts()
    return anomalies_df, anomaly_counter


def create_summary_with_corrections(anomalies_df, anomaly_counter, tab_type="radio"):
    """
    Construit un récap des anomalies avec le nombre de corrections proposées
    (en se basant sur la présence des colonnes 'Correction *' non vides).
    """
    summary_data = []
    correction_map = {}

    # Mapping 'type d'anomalie' -> colonne de correction
    if tab_type == "radio":
        correction_map = {
            "L'année de millésime n'est pas conforme": 'Correction Année',
            "Le diamètre n'est pas conforme": 'Correction Diamètre',
            'Incohérence Type Compteur': 'Correction Type Compteur',
            'SAPPEL: Incohérence Marque/Compteur (C)': 'Correction Marque',
            'SAPPEL: Incohérence Marque/Compteur (H)': 'Correction Marque',
            'Numéro de tête manquant': 'Correction Numéro de Tête',
            'KAMSTRUP: Protocole ≠ WMS': 'Correction Protocole Radio',
            'SAPPEL: Protocole ≠ OMS (année > 22)': 'Correction Protocole Radio',
            'SAPPEL: Protocole ≠ WMS (année <= 22)': 'Correction Protocole Radio',
        }
    elif tab_type == "tele":
        correction_map = {
            'Année millésime non conforme FP2E': 'Correction Année',
            'Diamètre non conforme FP2E': 'Correction Diamètre',
            'Incohérence Type Compteur': 'Correction Type Compteur',
            'SAPPEL: Incohérence Marque/Compteur (C)': 'Correction Marque',
            'SAPPEL: Incohérence Marque/Compteur (H)': 'Correction Marque',
            'Protocole incorrect (devrait être LRA)': 'Correction Protocole Radio',
            'Protocole incorrect (devrait être SGX)': 'Correction Protocole Radio',
        }
    elif tab_type == "manuelle":
        correction_map = {
            "L'année de millésime n'est pas conforme": 'Correction Année',
            "Le diamètre n'est pas conforme": 'Correction Diamètre',
            "L'année de millésime n'est pas conforme (FP2E+suffixe)": 'Correction Année',
            "Le diamètre n'est pas conforme (FP2E+suffixe)": 'Correction Diamètre',
            'SAPPEL: Incohérence Marque/Compteur (C)': 'Correction Marque',
            'SAPPEL: Incohérence Marque/Compteur (H)': 'Correction Marque',
            'ITRON: Incohérence Marque/Compteur': 'Correction Marque',
            'Incohérence Type Compteur': 'Correction Type Compteur',
        }

    # Alimente le récap : (type, occurrences, nb_lignes_avec_correction_proposée)
    for anomaly_type, count in anomaly_counter.items():
        correction_col = correction_map.get(anomaly_type)
        corrections_count = 0
        if correction_col:
            mask = anomalies_df['Anomalie'].str.contains(re.escape(anomaly_type), na=False) & (anomalies_df[correction_col] != '')
            corrections_count = anomalies_df[mask].shape[0]
        summary_data.append([anomaly_type, count, corrections_count])

    summary_df = pd.DataFrame(summary_data, columns=["Type d'anomalie", 'Nombre de cas', 'Corrections Proposées'])
    return summary_df


def creer_rapport_excel_detaille(output_path, anomalies_df, anomaly_counter, tab_type):
    """
    Génère un fichier Excel avec :
      - onglet 'Récapitulatif' (tableau synthèse + liens vers détails),
      - onglet 'Toutes_Anomalies' (toutes les lignes concernées),
      - 1 onglet par type d'anomalie (filtré), avec surlignage des colonnes liées.
    """
    summary_df = create_summary_with_corrections(anomalies_df, anomaly_counter, tab_type=tab_type)

    # Colonnes à mettre en évidence selon le type d'anomalie
    anomaly_columns_map = {}
    if tab_type == "radio":
        anomaly_columns_map = {
            "KAMSTRUP: Protocole ≠ WMS": ['Protocole Radio'],
            "SAPPEL: Protocole ≠ OMS (année > 22)": ['Protocole Radio'],
            "SAPPEL: Protocole ≠ WMS (année <= 22)": ['Protocole Radio'],
            "Marque manquante": ['Marque'],
            "Marque non autorisée en radiorelève": ['Marque'],
            "Numéro de compteur manquant": ['Numéro de compteur'],
            "Numéro de tête manquant": ['Numéro de tête'],
            "Coordonnées GPS non numériques": ['Latitude', 'Longitude'],
            "Coordonnées GPS invalides": ['Latitude', 'Longitude'],
            "Diamètre manquant": ['Diametre'],
            "Année de fabrication manquante": ['Année de fabrication'],
            "KAMSTRUP: Compteur ≠ 8 caractères": ['Numéro de compteur'],
            "KAMSTRUP: Compteur ≠ Tête": ['Numéro de compteur', 'Numéro de tête'],
            "KAMSTRUP: Compteur ou Tête non numérique": ['Numéro de compteur', 'Numéro de tête'],
            "KAMSTRUP: Diamètre hors plage": ['Diametre'],
            "KAMSTRUP: Format FP2E invalide": ['Numéro de compteur'],
            "U Kamstrup: Format FP2E invalide": ['Numéro de compteur'],
            "U Kamstrup: Tête ≠ 8 chiffres": ['Numéro de tête'],
            "SAPPEL: Tête DME ≠ 15 caractères": ['Numéro de tête'],
            "SAPPEL SEN4: Tête ≠ 16 caractères": ['Numéro de tête'],
            "SAPPEL: Compteur ne commence pas par C ou H": ['Numéro de compteur'],
            "SAPPEL: Incohérence Marque/Compteur (C)": ['Marque'],
            "SAPPEL: Incohérence Marque/Compteur (H)": ['Marque'],
            "ITRON: Compteur ne commence pas par I ou D": ['Numéro de compteur'],
            "Le numéro de compteur n'est pas conforme": ['Numéro de compteur'],
            "Le diamètre n'est pas conforme": ['Diametre'],
            "L'année de millésime n'est pas conforme": ['Année de fabrication'],
            "Incohérence Type Compteur": ['Type Compteur'],
            "Type Compteur non autorisé": ['Type Compteur'],
        }
    elif tab_type == "tele":
        anomaly_columns_map = {
            "Protocole incorrect (devrait être LRA)": ['Protocole Radio'],
            "Protocole incorrect (devrait être SGX)": ['Protocole Radio'],
            "Marque manquante": ['Marque'],
            "Marque non autorisée en télérelève": ['Marque'],
            "Numéro de compteur manquant": ['Numéro de compteur'],
            "Numéro de tête manquant": ['Numéro de tête'],
            "Coordonnées GPS non numériques": ['Latitude', 'Longitude'],
            "Coordonnées GPS invalides": ['Latitude', 'Longitude'],
            "Diamètre manquant": ['Diametre'],
            "Année de fabrication manquante": ['Année de fabrication'],
            "KAMSTRUP: Compteur ≠ 8 caractères": ['Numéro de compteur'],
            "KAMSTRUP: Compteur ≠ Tête": ['Numéro de compteur', 'Numéro de tête'],
            "KAMSTRUP: Compteur ou Tête non numérique": ['Numéro de compteur', 'Numéro de tête'],
            "KAMSTRUP: Diamètre hors de la plage [15, 400]": ['Diametre'],
            "KAMSTRUP: Format FP2E invalide": ['Numéro de compteur'],
            "U Kamstrup: Format FP2E invalide": ['Numéro de compteur'],
            "U Kamstrup: Tête ≠ 8 chiffres": ['Numéro de tête'],
            "SAPPEL: Tête ≠ 16 caractères": ['Numéro de tête'],
            "SAPPEL SEN3: Tête ≠ 15 caractères": ['Numéro de tête'],
            "SAPPEL: Incohérence Marque/Compteur (C)": ['Marque'],
            "SAPPEL: Incohérence Marque/Compteur (H)": ['Marque'],
            "ITRON: Tête ≠ 8 caractères": ['Numéro de tête'],
            'ITRON manuel: doit commencer par "I" ou "D"': ['Numéro de compteur'],
            'SAPPEL manuel: doit commencer par "C" ou "H"': ['Numéro de compteur'],
            "Format de compteur non FP2E": ['Numéro de compteur'],
            "Année millésime non conforme FP2E": ['Année de fabrication'],
            "Diamètre non conforme FP2E": ['Diametre'],
            "Incohérence Type Compteur": ['Type Compteur'],
            "Type Compteur non autorisé": ['Type Compteur'],
        }
    elif tab_type == "manuelle":
        anomaly_columns_map = {
            "Coordonnées GPS non numériques": ['Latitude', 'Longitude'],
            "Coordonnées GPS invalides": ['Latitude', 'Longitude'],
            "L'année de millésime n'est pas conforme": ['Année de fabrication'],
            "Le diamètre n'est pas conforme": ['Diametre'],
            "L'année de millésime n'est pas conforme (FP2E+suffixe)": ['Année de fabrication'],
            "Le diamètre n'est pas conforme (FP2E+suffixe)": ['Diametre'],
            "Le numéro de compteur n'est pas conforme (FP2E+suffixe)": ['Numéro de compteur'],
            "Compteur non-FP2E pour SAPPEL/ITRON": ['Numéro de compteur'],
            "SAPPEL: Incohérence Marque/Compteur (C)": ['Marque'],
            "SAPPEL: Incohérence Marque/Compteur (H)": ['Marque'],
            "ITRON: Incohérence Marque/Compteur": ['Marque'],
            "Incohérence Type Compteur": ['Type Compteur'],
            "Type Compteur non autorisé": ['Type Compteur'],
        }

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        wb = writer.book

        # Feuille récap
        ws_summary = wb.create_sheet(title="Récapitulatif", index=0)
        ws_summary['A1'] = "Récapitulatif des anomalies"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.append([])

        for r_idx, row_data in enumerate(dataframe_to_rows(summary_df, index=False, header=True)):
            ws_summary.append(row_data)

        # Feuille toutes anomalies
        anomalies_df.to_excel(writer, sheet_name="Toutes_Anomalies", index=False)
        ws_all_anomalies = wb["Toutes_Anomalies"]

        header_font = Font(bold=True)
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        for cell in ws_all_anomalies[1]:
            cell.font = header_font
        for cell in ws_summary[3]:
            cell.font = header_font

        # Surligne les colonnes concernées par chaque anomalie
        for row_num_all, df_row in enumerate(anomalies_df.iterrows(), 2):
            for anomaly in str(df_row[1]['Anomalie']).split(' / '):
                cols_a_surligner = anomaly_columns_map.get(anomaly.strip()) or colonnes_surlignage_defaut(anomaly)
                if cols_a_surligner:
                    for col_name in cols_a_surligner:
                        try:
                            ws_all_anomalies.cell(
                                row=row_num_all,
                                column=list(anomalies_df.columns).index(col_name) + 1
                            ).fill = red_fill
                        except ValueError:
                            pass

        # Ajuste largeur colonnes
        for col in ws_all_anomalies.columns:
            ws_all_anomalies.column_dimensions[get_column_letter(col[0].column)].width = max(
                len(str(cell.value)) for cell in col if cell.value
            ) + 2

        # Crée un onglet par type d'anomalie + liens depuis le récap
        created_sheet_names = {"Récapitulatif", "Toutes_Anomalies"}
        for idx, (anomaly_type, count, corrections) in enumerate(summary_df.values):
            current_row_num_in_summary = 4 + idx

            # Nom d'onglet safe (sans caractères interdits)
            sheet_name = re.sub(r'[\\/?*\[\]:()\'"<>|]', '', anomaly_type[:28]).replace(' ', '_').strip()
            original_sheet_name = sheet_name
            s_counter = 1
            while sheet_name in created_sheet_names:
                sheet_name = f"{original_sheet_name[:28]}_{s_counter}"
                s_counter += 1
            created_sheet_names.add(sheet_name)

            # Lien interne depuis le récap
            summary_cell = ws_summary.cell(row=current_row_num_in_summary, column=1)
            summary_cell.hyperlink = f"#'{sheet_name}'!A1"
            summary_cell.font = Font(underline="single", color="0563C1")

            # Feuille détaillée filtrée
            ws_detail = wb.create_sheet(title=sheet_name)
            filtered_df = anomalies_df[anomalies_df['Anomalie'].str.contains(re.escape(anomaly_type), na=False)]

            for r in dataframe_to_rows(filtered_df, index=False, header=True):
                ws_detail.append(r)

            for cell in ws_detail[1]:
                cell.font = header_font

            # Surlignage des colonnes liées dans l'onglet de détail
            for row_num_detail, df_row_detail in enumerate(filtered_df.iterrows(), 2):
                for anomaly in str(df_row_detail[1]['Anomalie']).split(' / '):
                    cols_a_surligner = anomaly_columns_map.get(anomaly.strip()) or colonnes_surlignage_defaut(anomaly)
                    if cols_a_surligner:
                        for col_name in cols_a_surligner:
                            try:
                                ws_detail.cell(
                                    row=row_num_detail,
                                    column=list(filtered_df.columns).index(col_name) + 1
                                ).fill = red_fill
                            except ValueError:
                                pass

            # Ajuste largeur colonnes
            for col in ws_detail.columns:
                ws_detail.column_dimensions[get_column_letter(col[0].column)].width = max(
                    len(str(cell.value)) for cell in col if cell.value
                ) + 2

        # Nettoie la feuille par défaut si elle existe
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
