import pandas as pd
from datetime import datetime
import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Ta logique existante (ne pas modifier)
import logique_controles

# --- CONFIG ---
DOSSIER_ENTREE = "C:/Users/kcostisor/Desktop/Automatisation/Fichier_A_Anlayser"
DOSSIER_SORTIE = "C:/Users/kcostisor/Desktop/Automatisation/Rapports_Anomalies"
FICHIER_RADIO = "radioreleve.xlsx"
FICHIER_TELE = "telereleve.xlsx"
FICHIER_MANUELLE = "manuelle.xlsx"
# --- FIN CONFIG ---

# Colonnes à supprimer des exports Traités si présentes
COLONNES_A_SUPPR_TRAITES = ["Origine de localisation", "LB_GPSORIGINE"]

# Mapping colonnes à surligner (union des 3 modes)
ANOMALY_COLUMNS_MAP_UNION = {
    # Protocole
    "KAMSTRUP: Protocole ≠ WMS": ["Protocole Radio"],
    "SAPPEL: Protocole ≠ OMS (année > 22)": ["Protocole Radio"],
    "SAPPEL: Protocole ≠ WMS (année <= 22)": ["Protocole Radio"],
    "Protocole incorrect (devrait être LRA)": ["Protocole Radio"],
    "Protocole incorrect (devrait être SGX)": ["Protocole Radio"],
    # Champs manquants / GPS
    "Marque manquante": ["Marque"],
    "Marque non autorisée en radiorelève": ["Marque"],
    "Marque non autorisée en télérelève": ["Marque"],
    "Numéro de compteur manquant": ["Numéro de compteur"],
    "Numéro de tête manquant": ["Numéro de tête"],
    "Coordonnées GPS non numériques": ["Latitude", "Longitude"],
    "Coordonnées GPS invalides": ["Latitude", "Longitude"],
    "Diamètre manquant": ["Diametre"],
    "Année de fabrication manquante": ["Année de fabrication"],
    # KAMSTRUP
    "KAMSTRUP: Compteur ≠ 8 caractères": ["Numéro de compteur"],
    "KAMSTRUP: Compteur ≠ Tête": ["Numéro de compteur", "Numéro de tête"],
    "KAMSTRUP: Compteur ou Tête non numérique": ["Numéro de compteur", "Numéro de tête"],
    "KAMSTRUP: Diamètre hors plage": ["Diametre"],
    "KAMSTRUP: Format FP2E invalide": ["Numéro de compteur"],
    "U Kamstrup: Format FP2E invalide": ["Numéro de compteur"],
    "U Kamstrup: Tête ≠ 8 chiffres": ["Numéro de tête"],
    # SAPPEL
    "SAPPEL: Tête DME ≠ 15 caractères": ["Numéro de tête"],
    "SAPPEL: Tête ≠ 16 caractères": ["Numéro de tête"],
    "SAPPEL SEN4: Tête ≠ 16 caractères": ["Numéro de tête"],
    "SAPPEL SEN3: Tête ≠ 15 caractères": ["Numéro de tête"],
    "SAPPEL: Compteur ne commence pas par C ou H": ["Numéro de compteur"],
    "SAPPEL: Incohérence Marque/Compteur (C)": ["Marque"],
    "SAPPEL: Incohérence Marque/Compteur (H)": ["Marque"],
    # ITRON
    "ITRON: Compteur ne commence pas par I ou D": ["Numéro de compteur"],
    "ITRON: Tête ≠ 8 caractères": ["Numéro de tête"],
    'ITRON manuel: doit commencer par "I" ou "D"': ["Numéro de compteur"],
    'SAPPEL manuel: doit commencer par "C" ou "H"': ["Numéro de compteur"],
    # FP2E / génériques
    "Le numéro de compteur n'est pas conforme": ["Numéro de compteur"],
    "Le diamètre n'est pas conforme": ["Diametre"],
    "L'année de millésime n'est pas conforme": ["Année de fabrication"],
    "Format de compteur non FP2E": ["Numéro de compteur"],
    "Année millésime non conforme FP2E": ["Année de fabrication"],
    "Diamètre non conforme FP2E": ["Diametre"],
    "Incohérence Type Compteur": ["Type Compteur"],
    "Type Compteur non autorisé": ["Type Compteur"],
    # NOUVEAU: FP2E avec suffixe (marques autres)
    "L'année de millésime n'est pas conforme (FP2E+suffixe)": ["Année de fabrication"],
    "Le diamètre n'est pas conforme (FP2E+suffixe)": ["Diametre"],
    "Le numéro de compteur n'est pas conforme (FP2E+suffixe)": ["Numéro de compteur"],
    # Divers
    "Compteur non-FP2E pour SAPPEL/ITRON": ["Numéro de compteur"],
    "Erreur de format interne": ["Numéro de compteur"]
}

RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
HEADER_FONT = Font(bold=True)
LINK_FONT = Font(underline="single", color="0563C1")

def traite_key(val: object) -> str:
    """Retourne les 3 premiers chiffres de 'Traité', sinon 'NON_RENSEIGNE'."""
    if pd.isna(val):
        return "NON_RENSEIGNE"
    s = str(val).strip()
    m = re.match(r"^\s*(\d{3})", s)
    return m.group(1) if m else "NON_RENSEIGNE"

def adjust_col_widths(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[letter].width = max_len + 2

def write_df(ws, df: pd.DataFrame, bold_header=True):
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    if bold_header:
        for c in ws[1]:
            c.font = HEADER_FONT
    adjust_col_widths(ws)

def highlight_anomaly_cells(ws, df):
    if "Anomalie" not in df.columns:
        return
    cols_idx = {col: i+1 for i, col in enumerate(df.columns)}
    # header gras déjà fait
    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        anomalies_list = str(row.get("Anomalie", "")).split(" / ")
        for anomaly in anomalies_list:
            a = anomaly.strip()
            if not a:
                continue
            cols = ANOMALY_COLUMNS_MAP_UNION.get(a) or logique_controles.colonnes_surlignage_defaut(a)
            for col_name in cols:
                if col_name in cols_idx:
                    ws.cell(row=r_idx, column=cols_idx[col_name]).fill = RED_FILL

def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Récap par type d'anomalie (Nombre de cas)."""
    if "Anomalie" not in df.columns or df.empty:
        return pd.DataFrame(columns=["Type d'anomalie", "Nombre de cas"])
    counts = (
        df["Anomalie"]
        .fillna("")
        .str.split(" / ")
        .explode()
        .str.strip()
        .replace("", pd.NA)
        .dropna()
        .value_counts()
        .rename_axis("Type d'anomalie")
        .reset_index(name="Nombre de cas")
    )
    return counts

def sanitize_sheet_name(name: str) -> str:
    name = re.sub(r'[\\/?*\[\]:\'"<>\|]', '', name).strip()
    if not name:
        name = "Sheet"
    return name[:28]

def create_excel_traite(chemin_out: str, df_all: pd.DataFrame):
    """
    Crée un Excel par Traité avec :
      - Récapitulatif (liens vers feuilles d'anomalies)
      - Toutes_Anomalies
      - 1 feuille par type d'anomalie (surlignée)
    """
    wb = Workbook()
    # Supprimer la feuille par défaut
    default_ws = wb.active
    wb.remove(default_ws)

    # Feuille Récap
    ws_summary = wb.create_sheet("Récapitulatif")
    ws_summary["A1"] = "Récapitulatif des anomalies"
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary.append([])

    # Feuille Toutes_Anomalies
    ws_all = wb.create_sheet("Toutes_Anomalies")

    # Nettoyage colonnes à supprimer si présentes
    cols_a_suppr = [c for c in COLONNES_A_SUPPR_TRAITES if c in df_all.columns]
    df_all = df_all.drop(columns=cols_a_suppr) if cols_a_suppr else df_all

    # Ecriture + surlignage
    write_df(ws_all, df_all)
    highlight_anomaly_cells(ws_all, df_all)

    # Récap (compte par type)
    summary_df = build_summary(df_all)
    # écriture sous A3
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws_summary.append(r)
    # style header recap
    if ws_summary.max_row >= 3:
        for c in ws_summary[3]:
            c.font = HEADER_FONT
    adjust_col_widths(ws_summary)

    # Feuilles par type d'anomalie + liens
    created_names = {"Récapitulatif", "Toutes_Anomalies"}
    for idx, row in summary_df.iterrows():
        anomaly = row["Type d'anomalie"]
        if not isinstance(anomaly, str) or not anomaly.strip():
            continue
        sheet_name = sanitize_sheet_name(anomaly.replace(" ", "_"))
        base = sheet_name
        k = 1
        while sheet_name in created_names:
            sheet_name = f"{base[:24]}_{k}"
            k += 1
        created_names.add(sheet_name)

        # lien depuis recap
        rec_row = 4 + idx  # A3 = header recap; data commence à la ligne 4
        cell = ws_summary.cell(row=rec_row, column=1)
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.font = LINK_FONT

        # feuille détail filtrée
        ws_detail = wb.create_sheet(sheet_name)
        df_filter = df_all[df_all["Anomalie"].str.contains(re.escape(anomaly), na=False)].copy()
        write_df(ws_detail, df_filter)
        highlight_anomaly_cells(ws_detail, df_filter)

    wb.save(chemin_out)

def main():
    date_actuelle = datetime.now()
    date_str = date_actuelle.strftime('%Y_%B')
    print(f"--- Lancement du rapport pour {date_str} ---")

    try:
        os.makedirs(DOSSIER_SORTIE, exist_ok=True)

        anomalies_tous_modes = []

        tasks = [
            {"nom": "Radioreleve", "fichier_in": FICHIER_RADIO, "fonction_check": logique_controles.check_data_radio, "tab_type": "radio"},
            {"nom": "Telereleve", "fichier_in": FICHIER_TELE, "fonction_check": logique_controles.check_data_tele, "tab_type": "tele"},
            {"nom": "Manuelle", "fichier_in": FICHIER_MANUELLE, "fonction_check": logique_controles.check_data_manuelle, "tab_type": "manuelle"},
        ]

        for task in tasks:
            print(f"\nAnalyse du fichier {task['fichier_in']}...")
            df = pd.read_excel(os.path.join(DOSSIER_ENTREE, task["fichier_in"]))
            df = df.iloc[:-2].copy()  # comme avant

            anomalies_df, anomaly_counter = task["fonction_check"](df)
            print(f"-> {len(anomalies_df)} anomalies trouvées.")

            if not anomalies_df.empty:
                # Rapport détaillé par mode (inchangé)
                nom_rapport = f"Rapport_{task['nom']}_{date_str}.xlsx"
                chemin_rapport_final = os.path.join(DOSSIER_SORTIE, nom_rapport)
                logique_controles.creer_rapport_excel_detaille(
                    chemin_rapport_final, anomalies_df, anomaly_counter, task["tab_type"]
                )
                print(f"Rapport détaillé sauvegardé : {chemin_rapport_final}")

                # Collecte pour regroupement Traités
                tmp = anomalies_df.copy()
                if "Traité" not in tmp.columns:
                    tmp["Traité"] = "NON_RENSEIGNE"
                tmp["_Source_Mode"] = task["nom"]  # interne, non exporté
                anomalies_tous_modes.append(tmp)

        # --- Fichiers par Traité (3 premiers chiffres) ---
        if anomalies_tous_modes:
            df_all = pd.concat(anomalies_tous_modes, ignore_index=True, sort=False)
            out_dir = os.path.join(DOSSIER_SORTIE, "Traites")
            os.makedirs(out_dir, exist_ok=True)

            def key_func(v): return traite_key(v)

            for code3, grp in df_all.groupby(df_all["Traité"].map(key_func), dropna=False):
                # enlever la colonne interne et les deux colonnes à supprimer (si présentes)
                drop_cols = ["_Source_Mode"] + [c for c in COLONNES_A_SUPPR_TRAITES if c in grp.columns]
                grp_to_save = grp.drop(columns=[c for c in drop_cols if c in grp.columns])

                # tri pour lisibilité
                sort_cols = [c for c in ["Traité", "Index original"] if c in grp_to_save.columns]
                if sort_cols:
                    grp_to_save = grp_to_save.sort_values(sort_cols)

                nom_fic = f"Anomalies_Traite_{code3}_{date_str}.xlsx"
                chemin_out = os.path.join(out_dir, nom_fic)

                # Création Excel complet (Récap + Toutes_Anomalies + feuilles par anomalie)
                create_excel_traite(chemin_out, grp_to_save)
                print(f"[Traité {code3}] {len(grp_to_save)} lignes -> {chemin_out}")
        else:
            print("\nAucune anomalie détectée. Aucun rapport à produire.")

        # Pas d'envoi d'email.

    except FileNotFoundError as e:
        print(f"\nERREUR : Fichier introuvable. Assurez-vous que '{os.path.basename(e.filename)}' est dans : {DOSSIER_ENTREE}")
    except Exception as e:
        print(f"\nUne erreur inattendue est survenue : {e}")

if __name__ == "__main__":
    main()
